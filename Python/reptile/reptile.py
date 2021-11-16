# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup  # 网页解析，获取数据
import re  # 正则表达式，进行文字匹配
# import urllib.request  # 制定URL，获取网页数据
# import urllib.error  # 制定URL，获取网页数据
import requests  # 制定URL，获取网页数据
# import xlwt  # 进行Excel操作
import openpyxl   # 进行Excel操作
import sqlite3  # 进行SQLite数据库操作

# 请求头 - header
headers = {
    # 伪装 浏览器进行 请求数据
    "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1"
}


# 得到指定一个URL的网页内容
def askGetUrl(url):
    response = requests.get(url=url, headers=headers)
    code = response.status_code  # 请求码
    reason = response.reason  # 请求返回的请求状态提示语
    html_str: str = ''
    if code == 200:
        print("请求成功", reason)
        html_str = response.content.decode(encoding='utf-8')
        # print(html_str)
    else:
        print("请求失败", reason, code)
    return html_str


# link的 正则表达式
find_link = re.compile(r'<a href="(.*?)">')   # 创建正则表达式对象，表示规则(字符串的模式)
# <img width="100" alt="肖申克的救赎" src="https://img2.doubanio.com/view/photo/s_ratio_poster/public/p480747492.webp" class="">
find_img = re.compile(r'<img .*? src="(.*?)" .*?/>', re.S)   # re.S让换行符包含在字符中
# 标题
find_title = re.compile(r'<span class="title">(.*?)</span>')
find_title_other = re.compile(r'<span class="other">(.*?)</span>')
# 导演等信息
find_daoyan = re.compile(r'<p class="">(.*?)</p>', re.S)
# 评分
find_rate = re.compile(r'<span class="rating_num" property="v:average">(.*)</span>')
# 多少人评价
find_num = re.compile(r'<span>(.*)人评价</span>')
# inq
find_inq = re.compile(r'<span class="inq">(.*)</span>')


# 抓取网页
def getData(url):
    # 获取到的所有数据
    dataList = []
    for i in range(0, 10):
        baseUrl = url+str(i*25)
        html_str = askGetUrl(url=baseUrl)
        # 逐一解析数据
        soup = BeautifulSoup(html_str, 'html.parser')
        # 查找符合要求的字符串，形成列表 <div class="item">
        for item in soup.find_all('div', class_='item'):
            # print(item)   # 查看电影item 的信息
            item = str(item)

            # 影片详情的链接
            link = ''
            links = re.findall(pattern=find_link, string=item)
            if len(links) > 0:
                link = links[0]

            # 影片的图片链接
            img = ''
            imgs = re.findall(pattern=find_img, string=item)
            if len(imgs) > 0:
                img = imgs[0]

            # 影片的名称
            title = ''
            titles = re.findall(pattern=find_title, string=item)
            for tit in titles:
                title += tit
            # 别名
            titles_others = re.findall(pattern=find_title_other, string=item)
            for tit in titles_others:
                title += tit
            # if title.find("海上钢琴师") >= 0:
            #     print(title)
            #     pass
            # print(type(title))

            # 导演的信息
            director = ''
            directors = re.findall(pattern=find_daoyan, string=item)
            if len(directors) > 0:
                director = directors[0]
                director = director.replace("\n", "").strip().replace("<br/>                            ", "\n")

            # 影片的评分
            rate = ''
            rates = re.findall(pattern=find_rate, string=item)
            if len(rates) > 0:
                rate = rates[0]

            # 评价人数
            num = ''
            nums = re.findall(pattern=find_num, string=item)
            if len(nums) > 0:
                num = nums[0]

            # inq
            inq = ''
            inqs = re.findall(pattern=find_inq, string=item)
            if len(inqs) > 0:
                inq = inqs[0]

            # dict1 = {
            #     "link": link,
            #     "img": img,
            #     "title": title,
            #     "director": director,
            #     "inq": inq,
            #     "rate": rate,
            #     "num": num,
            # }
            # dataList.append(dict1)

            # ['电影名', '电影图片', '详情地址', '导演信息', '评分', '评价人数', '简介']
            lists = [title, img, link, director, rate, num, inq]
            dataList.append(lists)
    return dataList


# 保存数据到Excel中
def saveDataToExcel(lists, path):
    print("保存数据到Excel中")
    # 1.创建Excel的操作对象
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(title="豆瓣电影Top250")
    # 第一行的标题
    xlsx_titles = ['电影名', '电影图片', '详情地址', '导演信息', '评分', '评价人数', '简介']
    for i in range(0, len(xlsx_titles)):
        worksheet.cell(row=1, column=i+1, value=xlsx_titles[i])

    for i in range(0, len(lists)):
        das = lists[i]
        for j in range(0, len(das)):
            worksheet.cell(row=2+i, column=j+1, value=das[j])
    workbook.save(filename=path)


# 使用SQLite进行保存数据
def saveDataToSQLite(lists, path):
    # 初始化 数据库
    # initSQLiteData(path=path)
    # return ""
    # 创建 或者 打开数据库
    db_conn = sqlite3.connect(database=path)
    # 获取游标
    db_cursor = db_conn.cursor()
    # 执行的sql语句
    for i in range(0, len(lists)):
        das = lists[i]
        # das.insert(0, i+1)
        for j in range(len(das)):
            # if j != 4:
            das[j] = '"' + das[j] + '"'
        db_sql = '''
            insert into doubanMovieTop250 (name, img, info_link, director, rate, evaluator, info)
            values (%s) ''' % ",".join(das)
        print(db_sql)
        # 开始执行
        db_cursor.execute(db_sql)
        # 提交
        db_conn.commit()
        # db_sql = """
        #     insert into doubanMovieTop250 (name, img, info_link, director, rate, evaluator, info)
        #     values ('%s', '%s', '%s', '%s', %0.1f, '%s', '%s')
        #     """ % (das[0], das[1], das[2], das[3], das[4], das[5], das[6])
    # db_sql = """
    #             insert into doubanMovieTop250 (name, img, info_link, director, rate, evaluator, info)
    #             values ('%s', '%s', '%s', '%s', %0.1f, '%s', '%s')
    #             """ % ('电影名1', '电影图片1', '详情地址1', '导演信息1', 9.8, '评价人数1', '简介1')
    # db_sql = """
    #             insert into doubanMovieTop250 (name, img, info_link, director, rate, evaluator, info)
    #             values ('电影名2', '电影图片1', '详情地址1', '导演信息1', 9.8, '评价人数1', '简介1')
    #             """

    # 关闭
    db_conn.close()
    # db_cursor.close()


# 开始初始化 数据库
def initSQLiteData(path):
    # 打开或者创建 数据库
    db = sqlite3.connect(database=path)
    # 游标
    db_c = db.cursor()
    # 执行的语句 - 创建 表 primary ['电影名', '电影图片', '详情地址', '导演信息', '评分', '评价人数', '简介']
    sql1 = """
        create table doubanMovieTop250
        (
          id integer primary key autoincrement,
          name text,
          img text,
          info_link text,
          director text,
          rate real,
          evaluator text,
          info text
        )
    """
    # 开始执行
    db_c.execute(sql1)
    # 提交
    db.commit()
    # 关闭
    db_c.close()
    db.close()


# 主函数 进入
def main():
    # 请求的基础的URL
    base_url = "https://movie.douban.com/top250?start="
    # 1. 抓取数据 2.解析数据
    lists1 = getData(url=base_url)
    # print(lists1)
    # 3.0 保存数据
    # 3.1 通过Excel 保存数据
    # save_path = "豆瓣电影Top250.xlsx"
    # saveDataToExcel(lists=lists1, path=save_path)

    # 3.2 通过sqlite 保存数据
    db_path = "doubanMovieTop250.db"
    saveDataToSQLite(lists=lists1, path=db_path)

if __name__ == "__main__":
    main()   # 调用函数 进行运行程序
    print("数据已经全部保存完毕，请查看~")
"""
# 请求地址
url_path = base_url + "0"
# 发起请求
# data = requests.get(url=url_path, headers=headers)
# print(type(data))
# print(data.json())
# 解析数据
content = askGetUrl(url=url_path)
print(content)
# requests.models.Response
"""

"""
BeautifulSoup4将复杂的HTML文档换成一个复杂的树形结构，每个节点都是Python对象，所有的对象归纳为4种：
- tag 标签及其内容，拿到它所找到的第一个内容
- NavigableString .string标签里面的内容-字符串 .attrs 标签的属性
- BeautifulSoup 表示整个文档
- Comment 是一个特殊的NavigableString， 输出内容不包含注释符号
"""

# 文档的遍历
# bs.head.contents

"""
文档的搜索
1. find_all 字符串过滤：会查找与字符串完全匹配的内容

2. 正则表达式搜索：使用search() 方法来匹配内容
"""


# 测试Excel
# workbook = openpyxl.Workbook()
# worksheet = workbook.create_sheet(title="Sheet")
# # worksheet["A1"] = "hello"
# worksheet.cell(row=1, column=1, value='Hello')
# workbook.save(filename="./test1.xlsx")

# 九九乘法表
# workbook = openpyxl.Workbook()
# worksheet = workbook.create_sheet(title="九九乘法表")
# for i in range(1, 10):
#     for j in range(1, i+1):
#         value = '%d*%d=%d' % (j, i, i*j)
#         worksheet.cell(row=i, column=j, value=value)
# workbook.save(filename="./test1.xlsx")


# 去除开头的换行符 和 空格
# str1 = """
#                             导演: 弗兰克·德拉邦特 Frank Darabont   主演: 蒂姆·罗宾斯 Tim Robbins /...<br/>
#                             1994 / 美国 / 犯罪 剧情
#
# """
#
# str2 = str1.replace("\n", "")
#
# str3 = str2.strip()
# print(str1)


# sqlite3 的练习操作

# con = sqlite3.connect(database="test.db")   # 打开或者创建数据库
# print("数据库打开成功了")
#
# c = con.cursor()   # 获取游标
#
# # 执行的sql语句
# sql = """
#     create table company
#         (id int primary key not null,
#         name text not null,
#         age int not null,
#         address char(50),
#         salary real);
# """
#
# # 执行sql语句
# c.execute(sql)
# con.commit()   # 提交数据库操作
# con.close()    # 关闭数据库连接
#
# print("成功建表")


# 插入数据

# conn = sqlite3.connect(database="test.db")   # 打开或者创建数据库
# c1 = conn.cursor()  # 获取游标
#
# # 需要执行的语句
# sql1 = """
#      insert into company (id, name, age, address, salary)
#      values (1, '张三', 43, "北京", 18000);
# """
#
# sql2 = """
#      insert into company (id, name, age, address, salary)
#      values (2, "李四", 34, "上海", 30000);
# """
#
# # 执行的sql语句
# c1.execute(sql1)
# c1.execute(sql2)
# # 对执行的语句进行提交
# conn.commit()
# # 关闭数据库
# conn.close()


# 执行查询语句

# # 打开 或者 创建 数据库
# conn_search = sqlite3.connect(database="test.db")
# # 获取游标
# c_search = conn_search.cursor()
#
# # 执行的sql语句
# sql_search = """
#     select id, name, age, address, salary from company
# """
#
# # 执行
# lists = c_search.execute(sql_search)
# for item in lists:
#     print(item)
#
# # 提交
# # conn_search.commit()
# # 关闭数据库
# conn_search.close()
# print("完成")


