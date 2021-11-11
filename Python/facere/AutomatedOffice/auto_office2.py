"""
 Excel 公式、图表以及单元格的操作
"""

from openpyxl import Workbook
# from openpyxl import Workbook
from openpyxl import worksheet

# 文件路径
path = 'source/sum.xlsx'

# Excel的公式
# SUM 求和公式函数
wb = Workbook()
ws = wb.active
ws['A1'] = 1000
ws['A2'] = 900
ws['A3'] = 890
ws['A4'] = '=sum(A1:A3)'

# VLOOkUP

ws.row_dimensions[1].height = 80
ws.unmerge_cells()

wb.save(filename=path)
