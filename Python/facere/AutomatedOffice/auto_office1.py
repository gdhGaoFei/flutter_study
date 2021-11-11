"""
 Excel文档的写入。增删改查
"""

import openpyxl
from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.utils import get_column_letter # 字母到数字转化
from openpyxl.utils import column_index_from_string # 数字到字母之间的转化

path = "source/6377204931816695924077953.xlsx"

# 获取Excel
wb = openpyxl.load_workbook(path)
print(wb)

# 所有的工作薄表的名称
sheetnames = wb.sheetnames
print(sheetnames)

# 在工作簿中获取工作表
ws_sheet1 = wb['Sheet1']
print(type(ws_sheet1))
print(ws_sheet1.title)

# 有多少个列
max_row = ws_sheet1.max_row
print(max_row)
# 有多少个行
max_column = ws_sheet1.max_column
print(max_column)

# 获取某个单元格
a1 = ws_sheet1['A1']
print(a1)
# print(a1.value)
print(a1.value)
for i in range(0, max_column):
    row = ws_sheet1.cell(row=2, column=i+1)
    # ws_sheet1.cell(r)
    print(row.value)

print(get_column_letter(45))
print(column_index_from_string("DF"))

# 从工作表获取列和行
cells1 = ws_sheet1["A1":"D16"]
print(cells1)
for cells in cells1:
    for cell in cells:
        print(cell.value)

list_columns = list(ws_sheet1.columns)
print(list_columns)
for lcs in list_columns:
    for lc in lcs:
        print(lc.value)

ws_sheet1["E2"] = "时间"
wb.save(filename=path)

# 插入一行或者多行 删除多行或者一行
# ws_sheet1.insert(index=5)
# ws_sheet1.insert_rows(idx=5)
# ws_sheet1.insert_rows(idx=7, amount=3)
# wb.save(filename=path)
# ws_sheet1.delete_rows(idx=7, amount=3)
# ws_sheet1.delete_rows(idx=5)
# wb.save(filename=path)

# 插入一列或者多列 删除多列或者一列
# ws_sheet1.insert_cols(idx=3)
# ws_sheet1.insert_cols(idx=3, amount=3)
# ws_sheet1.delete_cols(idx=3, amount=3)
# ws_sheet1.delete_cols(idx=3)
# wb.save(filename=path)

