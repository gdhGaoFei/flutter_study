"""
自动化办公：Excel的文档创建、工作簿、工作表的创建、删除、查看
"""

import openpyxl
from openpyxl import Workbook

path = 'source/6377204931816695924077953.xlsx'
data = openpyxl.open(filename=path) # xlrd.open_workbook(filename=path)

# 创建的对象 - 工作簿
wb = Workbook()

# 保存一个Excel文件 需要是绝对的路径 /Users/gaodehua/Desktop/MineProject/Study/Python/facere/AutomatedOffice/
path1 = "source/test.xlsx"
# wb.save(filename=path1)
#
#
# # 在这个Excel文件中创建 sheet
# wb.create_sheet(title="MySheet")
# wb.save(filename=path1)
#
# for i in range(3):
#     wb.create_sheet(title="MySheet{}".format(i+1), index=i)
# wb.save(filename=path1)
#
# # 查看工作表的名字
# sheet_names = wb.sheetnames
# print(sheet_names)
#
# # 改变工作表的名字 先获取工作表 之后进行赋值 再保存即可
# ws_name = wb['Sheet']
# ws_name.title = "MySheet{}".format(len(sheet_names))
# wb.save(filename=path1)
# print(wb.sheetnames)

data1 = openpyxl.open(filename=path1)
# 获取活跃表 ====
# print(wb.sheetnames)
ws_active = data1.active# wb.active
print(ws_active.title)

# 工作表的删除
print(data1.sheetnames)
del data1['MySheet']
print(data1.sheetnames)
data1.save(filename=path1)
# pass